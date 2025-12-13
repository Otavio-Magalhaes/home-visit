import csv
from datetime import date, datetime, timedelta
import io
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from sqlalchemy.orm import Session
from sqlalchemy import func, cast, case, distinct, and_
from geoalchemy2 import Geography
from geoalchemy2.elements import WKTElement

from app.models.Residence import Residence
from app.models.Resident import Resident
from app.models.HealthSituation import HealthSituation
from app.schemas.dashboard_schema import DashboardFilterSchema

def _apply_dashboard_filters(db: Session, filters: DashboardFilterSchema):  
    center_point = WKTElement(f'POINT({filters.longitude} {filters.latitude})', srid=4326)

    # Inicia a query focada no Morador, trazendo a casa e a situação de saúde
    query = db.query(Resident)\
        .join(Residence, Resident.residence_id == Residence.id)\
        .outerjoin(HealthSituation, Resident.id == HealthSituation.resident_id)

    # 1. Filtro Geográfico
    query = query.filter(
        func.ST_DWithin(
            cast(Residence.geo_location, Geography), 
            cast(center_point, Geography), 
            filters.radius_meters
        )
    )

    # ==========================================
    # 2. Filtros de MORADOR
    # ==========================================
    if filters.sexo:
        query = query.filter(Resident.sexo == filters.sexo)
    if filters.raca_cor:
        query = query.filter(Resident.raca_cor == filters.raca_cor)
    if filters.escolaridade:
        query = query.filter(Resident.escolaridade == filters.escolaridade)
    if filters.situacao_mercado:
        query = query.filter(Resident.situacao_mercado == filters.situacao_mercado)
    if filters.frequenta_escola is not None:
        query = query.filter(Resident.frequenta_escola == filters.frequenta_escola)
    if filters.possui_deficiencia is not None:
        query = query.filter(Resident.possui_deficiencia == filters.possui_deficiencia)
    if filters.eh_responsavel_familiar is not None:
        query = query.filter(Resident.eh_responsavel_familiar == filters.eh_responsavel_familiar)
    
    # --- Novos campos de Morador ---
    if filters.orientacao_sexual:
        query = query.filter(Resident.orientacao_sexual == filters.orientacao_sexual)
    if filters.identidade_genero:
        query = query.filter(Resident.identidade_genero == filters.identidade_genero)
    if filters.membro_povo_tradicional is not None:
        query = query.filter(Resident.membro_povo_tradicional == filters.membro_povo_tradicional)
    if filters.possui_plano_saude is not None:
        query = query.filter(Resident.possui_plano_saude == filters.possui_plano_saude)
    if filters.participa_grupos_comunitarios is not None:
        query = query.filter(Resident.participa_grupos_comunitarios == filters.participa_grupos_comunitarios)

    # Lógica de Idade
    if filters.idade_minima:
        dt_limite = date.today() - timedelta(days=filters.idade_minima * 365.25)
        query = query.filter(Resident.data_nascimento <= dt_limite)
    
    if filters.idade_maxima:
        dt_limite = date.today() - timedelta(days=filters.idade_maxima * 365.25)
        query = query.filter(Resident.data_nascimento >= dt_limite)

    # ==========================================
    # 3. Filtros de RESIDÊNCIA
    # ==========================================
    if filters.microarea:
        query = query.filter(Residence.microarea == filters.microarea)
    if filters.tipo_imovel:
        query = query.filter(Residence.tipo_imovel == filters.tipo_imovel)
    if filters.situacao_moradia:
        query = query.filter(Residence.situacao_moradia == filters.situacao_moradia)
    if filters.abastecimento_agua:
        query = query.filter(Residence.abastecimento_agua == filters.abastecimento_agua)
    
    # --- Novo campo de Residência ---
    if filters.tratamento_agua:
        query = query.filter(Residence.tratamento_agua == filters.tratamento_agua)

    if filters.escoamento_sanitario:
        query = query.filter(Residence.escoamento_sanitario == filters.escoamento_sanitario)
    if filters.destino_lixo:
        query = query.filter(Residence.destino_lixo == filters.destino_lixo)
    if filters.material_paredes:
        query = query.filter(Residence.material_paredes == filters.material_paredes)
    if filters.disponibilidade_energia is not None:
        query = query.filter(Residence.disponibilidade_energia == filters.disponibilidade_energia)
    if filters.possui_animais is not None:
        query = query.filter(Residence.possui_animais == filters.possui_animais)

    # ==========================================
    # 4. Filtros de SAÚDE (HealthSituation)
    # ==========================================
    if filters.esta_gestante is not None:
        query = query.filter(HealthSituation.esta_gestante == filters.esta_gestante)
    if filters.esta_acamado is not None:
        query = query.filter(HealthSituation.esta_acamado == filters.esta_acamado)
    if filters.esta_domiciliado is not None:
        query = query.filter(HealthSituation.esta_domiciliado == filters.esta_domiciliado)
    
    # --- Novos campos de Saúde ---
    if filters.peso_situacao:
        query = query.filter(HealthSituation.peso_situacao == filters.peso_situacao)
    
    # Doenças
    if filters.tem_hipertensao is not None:
        query = query.filter(HealthSituation.tem_hipertensao == filters.tem_hipertensao)
    if filters.tem_diabetes is not None:
        query = query.filter(HealthSituation.tem_diabetes == filters.tem_diabetes)
    if filters.tem_doenca_respiratoria is not None:
        query = query.filter(HealthSituation.tem_doenca_respiratoria == filters.tem_doenca_respiratoria)
    if filters.tem_doenca_cardiaca is not None:
        query = query.filter(HealthSituation.tem_doenca_cardiaca == filters.tem_doenca_cardiaca)
    if filters.tem_problemas_rins is not None:
        query = query.filter(HealthSituation.tem_problemas_rins == filters.tem_problemas_rins)
    if filters.tem_hanseniase is not None:
        query = query.filter(HealthSituation.tem_hanseniase == filters.tem_hanseniase)
    if filters.tem_tuberculose is not None:
        query = query.filter(HealthSituation.tem_tuberculose == filters.tem_tuberculose)
    if filters.tem_cancer is not None:
        query = query.filter(HealthSituation.tem_cancer == filters.tem_cancer)
    
    if filters.teve_avc_derrame is not None:
        query = query.filter(HealthSituation.teve_avc_derrame == filters.teve_avc_derrame)
    if filters.teve_infarto is not None:
        query = query.filter(HealthSituation.teve_infarto == filters.teve_infarto)

    # Outros riscos e internação
    if filters.internacao_ultimos_12_meses is not None:
        query = query.filter(HealthSituation.internacao_ultimos_12_meses == filters.internacao_ultimos_12_meses)
    if filters.diagnostico_problema_mental is not None:
        query = query.filter(HealthSituation.diagnostico_problema_mental == filters.diagnostico_problema_mental)
    if filters.usa_alcool is not None:
        query = query.filter(HealthSituation.usa_alcool == filters.usa_alcool)
    if filters.usa_outras_drogas is not None:
        query = query.filter(HealthSituation.usa_outras_drogas == filters.usa_outras_drogas)
    if filters.esta_fumante is not None:
        query = query.filter(HealthSituation.esta_fumante == filters.esta_fumante)
    
    if filters.usa_plantas_medicinais is not None:
        query = query.filter(HealthSituation.usa_plantas_medicinais == filters.usa_plantas_medicinais)
    if filters.usa_praticas_integrativas is not None:
        query = query.filter(HealthSituation.usa_praticas_integrativas == filters.usa_praticas_integrativas)

    if filters.em_situacao_rua is not None:
        query = query.filter(HealthSituation.em_situacao_rua == filters.em_situacao_rua)
    if filters.recebe_beneficio is not None:
        query = query.filter(HealthSituation.recebe_beneficio == filters.recebe_beneficio)
   
    if filters.tem_acesso_higiene_pessoal is not None:
        query = query.filter(HealthSituation.tem_acesso_higiene_pessoal == filters.tem_acesso_higiene_pessoal)
    if filters.possui_referencia_familiar is not None:
        query = query.filter(HealthSituation.possui_referencia_familiar == filters.possui_referencia_familiar)

    if filters.frequencia_alimentacao:
        query = query.filter(HealthSituation.frequencia_alimentacao == filters.frequencia_alimentacao)

    return query

def get_dashboard_statistics(db: Session, filters: DashboardFilterSchema):
  base_query = _apply_dashboard_filters(db, filters)

  stats = base_query.with_entities(
      func.count(Resident.id).label("total_pessoas"),
      func.count(distinct(Residence.id)).label("total_domicilios"),
      
   
      func.sum(
          case((func.extract('year', func.age(Resident.data_nascimento)) < 18, 1), else_=0)
      ).label("qtd_menores"),
      
      func.sum(
          case((func.extract('year', func.age(Resident.data_nascimento)) >= 18, 1), else_=0)
      ).label("qtd_maiores"),
      
   
  ).first()

  return {
      "filtros_aplicados": filters.dict(exclude_none=True),
      "resultados": {
          "total_pessoas": stats.total_pessoas or 0,
          "total_domicilios": stats.total_domicilios or 0,
          "qtd_maiores": stats.qtd_maiores or 0, # Novo campo
          "qtd_menores": stats.qtd_menores or 0, # Novo campo
      }
  }
def get_dashboard_map_pins(db: Session, filters: DashboardFilterSchema):
    """
    Retorna a lista de residências para plotar no mapa.
    """
    base_query = _apply_dashboard_filters(db, filters)

   
    pins = base_query.with_entities(
        Residence.id,
        func.ST_Y(Residence.geo_location).label("latitude"),  # Extrai Latitude (Y)
        func.ST_X(Residence.geo_location).label("longitude"), # Extrai Longitude (X)
        Residence.nome_logradouro,
        Residence.numero,
        Residence.bairro
    ).group_by(Residence.id).all()

    return [
        {
            "id": str(pin.id), 
            "lat": pin.latitude,  
            "lng": pin.longitude, 
            "titulo": f"{pin.nome_logradouro}, {pin.numero} - {pin.bairro}"
        }
        for pin in pins
    ]
def _get_model_columns(model):
    """
    Retorna nomes das colunas, removendo campos indesejados (binários)
    e garantindo que propriedades úteis sejam incluídas.
    """
    columns = [c.name for c in model.__table__.columns if c.name != 'geo_location']
    
    if model.__name__ == 'Residence':
        columns.extend(['latitude', 'longitude'])
        
    return columns
def _format_value(value):
    """Formata valores para ficar legível no Excel"""
    if value is None:
        return ""
    if hasattr(value, "value"): 
        return value.value
        
    if isinstance(value, datetime):
        return value.replace(tzinfo=None)

    if isinstance(value, list):
        return ", ".join(map(str, value))
        
    return value

def _style_and_adjust_worksheet(ws, headers):
    """Aplica estilo ao cabeçalho e ajusta a largura das colunas."""
    header_font = Font(bold=True, color="FFFFFF")
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    header_fill = PatternFill(start_color="4d7cc3", end_color="4d7cc3", fill_type="solid")

    for cell in ws[1]:
        cell.font = header_font
        cell.alignment = header_alignment
        cell.fill = header_fill

    for i, column_header in enumerate(headers, 1):
        column_letter = ws.cell(row=1, column=i).column_letter
        max_length = len(str(column_header))
        for cell in ws[column_letter]:
            if cell.value and len(str(cell.value)) > max_length:
                max_length = len(str(cell.value))
        adjusted_width = (max_length + 2) * 1.2
        ws.column_dimensions[column_letter].width = adjusted_width
        
def export_dashboard_data_xlsx(db: Session, filters: DashboardFilterSchema):
    # 1. Busca os dados filtrados
    base_query = _apply_dashboard_filters(db, filters)
    residents = base_query.all()

    # 2. Separa em dicionários para evitar duplicidade nas abas
    unique_residences = {}
    unique_health_situations = {}
    data_residents = []

    for resident in residents:
        data_residents.append(resident)
        
        if resident.residence:
            unique_residences[resident.residence.id] = resident.residence
            
        if resident.health_situation:
            unique_health_situations[resident.health_situation.id] = resident.health_situation

    wb = Workbook()

    # ==============================
    # ABA 1: RESIDÊNCIAS
    # ==============================
    ws1 = wb.active
    ws1.title = "Residencias"
    cols_residence = _get_model_columns(Residence)
    ws1.append(cols_residence)
    
    for res in unique_residences.values():
        row = []
        for col in cols_residence:
            # Tratamento especial: tenta pegar o atributo ou property
            # Isso funciona tanto para colunas normais quanto para @property (latitude/longitude)
            val = getattr(res, col, None)
            row.append(_format_value(val))
        ws1.append(row)

    _style_and_adjust_worksheet(ws1, cols_residence)

    # ==============================
    # ABA 2: MORADORES
    # ==============================
    ws2 = wb.create_sheet(title="Moradores")
    cols_resident = _get_model_columns(Resident)
    ws2.append(cols_resident)
    
    for r in data_residents:
        row = [_format_value(getattr(r, col, None)) for col in cols_resident]
        ws2.append(row)

    _style_and_adjust_worksheet(ws2, cols_resident)

    # ==============================
    # ABA 3: SAÚDE
    # ==============================
    ws3 = wb.create_sheet(title="Saude")
    cols_health = _get_model_columns(HealthSituation)
    ws3.append(cols_health)
    
    for h in unique_health_situations.values():
        row = [_format_value(getattr(h, col, None)) for col in cols_health]
        ws3.append(row)

    _style_and_adjust_worksheet(ws3, cols_health)

    # Finaliza e retorna o arquivo em memória
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output